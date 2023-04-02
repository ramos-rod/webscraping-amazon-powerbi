import requests
from bs4 import BeautifulSoup
import openpyxl
import time

# determinar URLs de pesquisa
amd_url = 'https://www.amazon.com/s?k=AMD+CPU&crid=3HJ59IY8491TB&sprefix=amd+c%2Caps%2C332&ref=nb_sb_noss_2'
intel_url = 'https://www.amazon.com/s?k=Intel+CPU&crid=HMGM4ADKO6IK&sprefix=intel+cp%2Caps%2C281&ref=nb_sb_noss_2'

# determinar headers - imitar um navegador para fazer requests
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
           "Accept-Encoding": "gzip, deflate",
           "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
           "DNT": "1", "Connection": "close",
           "Upgrade-Insecure-Requests": "1"}

# lista para armazenar dados raspados
results = []

# improviso para extrair corretamente o clock max de cpus intel
keyword = 'up to'

# função raspar página
def scrape_page(url):
    # request da busca
    page = requests.get(url, headers=headers)
    # tempo de sobra para carregar a página (2s causou inconsistências)
    time.sleep(2)
    # extração do html da busca
    soup = BeautifulSoup(page.content, 'html.parser')
    # separação de itens listados no código através de tags
    items = soup.find_all('div', {'data-component-type': 's-search-result'})
    # iterando entre as listagens nos resultados de cada busca
    for item in items[:10]:
        # extração do url da listagem
        listing_url = 'https://www.amazon.com' + item.find('a', {'class': 'a-link-normal s-no-outline'})['href']
        # request de cada listagem
        listing_page = requests.get(listing_url, headers=headers)
        # extração do html da listagem
        listing_soup = BeautifulSoup(listing_page.content, 'html.parser')
        # extração do modelo da cpu
        title = listing_soup.find(id='productTitle').get_text().strip()
        # improviso para não raspar outros produtos que contenham 'AMD' ou 'Intel' no título
        if 'AMD' in title[:3]:
            brand = 'AMD'
        elif 'Intel' in title[:5]:
            brand = 'Intel'
        else:
            continue
        time.sleep(1)
        # extração do preço
        price = listing_soup.find('span', {'class': 'a-offscreen'}).get_text().strip()
        # caso produto indisponível / price == None
        if price is None:
            price = 0
        # separação de itens do html da listagem por tag e classe (wattage)
        listing_items1 = listing_soup.find_all('tr', {'class': 'a-spacing-small po-wattage'})
        # iterando entre subclasses
        for item2 in listing_items1:
            # extração da potência em watts
            wattage = item2.find('span', {'class': 'a-size-base po-break-word'}).text.strip()
        # improviso intel - extração do clock speed pelo título por conta de inconsistências na listagem
        if keyword in title:
            # manipulação de string
            max_clock_speed = title[title.find('up to') + 6: title.find('up to') + 9] + ' GHz'
        # amd - extração do clock speed da descrição da listagem
        else:
            # separação de itens do html da listagem por tag e classe (cpu model speed)
            listing_items2 = listing_soup.find_all('tr', {'class': 'a-spacing-small po-cpu_model.speed'})
            # iterando entre subclasses
            for item3 in listing_items2:
                # extração da velocidade de clock máxima
                max_clock_speed = item3.find('span', {'class': 'a-size-base po-break-word'}).text.strip()
        # separação e itens do html da listagem por tag e classe (nota/rating)
        listing_items3 = listing_soup.find_all('i', {'class': 'a-icon a-icon-star a-star-5'})
        # iterando entre subclasses
        for item4 in listing_items3:
            # extração da nota/rating
            rating = item4.find('span', {'class': 'a-icon-alt'})
            if rating is not None:
                rating = rating.get_text().strip()
                break
            # notei que quando a nota está abaixo de 4.8, a classe da tag 'i' troca para o final a-star-4-5
            # nesse caso, se rating == 'None', significa que não achamos a classe a-star-5 nem a-icon-alt, então
            # troca-se a procura find_all para a classe que comporta entre 4 e 5 estrelas. ***melhorar no futuro***
            else:
                listing_items3 = listing_soup.find_all('i', {'class': 'a-icon a-icon-star a-star-4-5'})
                for item4 in listing_items3:
                    # extração da nota/rating
                    rating = item4.find('span', {'class': 'a-icon-alt'})
                    if rating is not None:
                        rating = rating.get_text().strip()
                        break
        # adicionar dados à lista
        results.append((brand, title, price, rating, max_clock_speed, wattage, listing_url))


scrape_page(intel_url)
scrape_page(amd_url)


print(f'{len(results)} results found for AMD and Intel CPUs')

# não operar excel caso 0 resultados
if len(results) == 0:
    exit()
else:
    workbook = openpyxl.Workbook()         # iniciar workbook excel
    worksheet = workbook.active            # iniciar pag1
    worksheet.cell(1, 1, 'Marca')          # determinando nome das colunas
    worksheet.cell(1, 2, 'Modelo')
    worksheet.cell(1, 3, 'Preço')
    worksheet.cell(1, 4, 'Avaliação')
    worksheet.cell(1, 5, 'Clock máx.(GHz)')
    worksheet.cell(1, 6, 'Potência máx.(watts)')
    worksheet.cell(1, 7, 'URL')
    for i, result in enumerate(results):   # iterando entre resultados e adicionando às respectivas células
        worksheet.cell(i+2, 1, result[0])
        worksheet.cell(i+2, 2, result[1])
        worksheet.cell(i+2, 3, result[2])
        worksheet.cell(i+2, 4, result[3])
        worksheet.cell(i+2, 5, result[4])
        worksheet.cell(i+2, 6, result[5])
        worksheet.cell(i+2, 7, result[6])

    workbook.save('C:\\Desktop\\Estudos Data\\scrapedtest.xlsx')  # salvando arquivo
